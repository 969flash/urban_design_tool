import Rhino
import Rhino.Geometry as rg
import ghpythonlib.components as ghcomp
import scriptcontext as sc
import os
from datetime import datetime
import sys
import System
from System import Activator, Type
from System.Runtime.InteropServices import Marshal


class Block:
    def __init__(self, region, landuse_name, block_id):
        self.region = region
        self.landuse_name = landuse_name
        self.block_id = block_id
        self.buildings = []  # 2D 건물라인 리스트


try:
    # optional: openpyxl fallback for saving without Excel COM or without a template
    from openpyxl import Workbook
    from openpyxl.styles import Font, Border, Side, Alignment
    from openpyxl.utils import get_column_letter

    OPENPYXL_AVAILABLE = True
except Exception:
    OPENPYXL_AVAILABLE = False

# Python 3 호환: basestring 별칭
if sys.version_info[0] >= 3:
    basestring = str


def extrude_srf(srf, height):
    """Surface or Brep를 z축 방향으로 height만큼 Extrude한 Brep 반환"""
    extrusion = ghcomp.Extrude(srf, rg.Vector3d(0, 0, height))
    return ghcomp.CapHoles(extrusion)


def get_layer_surfaces(doc, parent_name):
    """
    Landuse 하위 레이어별 Surface/Brep 수집 → dict로 반환
    ex: {'Commercial': [Brep1, Brep2], 'Residential': [Brep3, ...]}
    """
    layer_dict = {}
    for layer in doc.Layers:
        if not layer.IsVisible:
            continue

        # 부모-자식 레이어 관계 확인 (FullPath 기준)
        full_path = getattr(layer, "FullPath", layer.Name)
        parts = full_path.split("::") if full_path else []
        if len(parts) >= 2 and parts[0] == parent_name:
            # 직계 하위명 (Parent::Child[::Grand...])에서 Child 추출
            child_name = parts[1]

            objs = [
                obj
                for obj in doc.Objects.FindByLayer(layer)
                if isinstance(obj.Geometry, rg.Surface)
                or isinstance(obj.Geometry, rg.Brep)
            ]

            srfs = []
            for obj in objs:
                geo = obj.Geometry
                if isinstance(geo, rg.Surface):
                    srfs.append(rg.Brep.CreateFromSurface(geo))
                elif isinstance(geo, rg.Brep):
                    srfs.append(geo)

            # 동일 Child 이름으로 누적 (손자 레이어까지 포함 시 합쳐짐)
            if child_name in layer_dict:
                layer_dict[child_name].extend(srfs)
            else:
                layer_dict[child_name] = srfs

    return layer_dict


def is_point_on_srf(pt, srf):
    pt_on_srf = ghcomp.SurfaceClosestPoint(pt, srf).point
    return pt_on_srf.DistanceTo(pt) < 0.01


def get_point_inside_face(surface):

    # Brep에서 메시 생성 (기본 세분화 파라미터 사용)
    meshes = rg.Mesh.CreateFromBrep(surface, rg.MeshingParameters.Default)

    # 메시 리스트 중 첫 번째 메시 반환 (대부분 하나의 메시 생성됨)
    if meshes and len(meshes) > 0:
        mesh = meshes[0]
    # 첫 번째 삼각형(face) 인덱스의 세 점을 가져옴
    indices = mesh.Faces[0]

    p0 = mesh.Vertices[indices.A]
    p1 = mesh.Vertices[indices.B]
    p2 = mesh.Vertices[indices.C]

    # 삼각형 무게중심 계산
    return rg.Point3d(
        (p0.X + p1.X + p2.X) / 3.0,
        (p0.Y + p1.Y + p2.Y) / 3.0,
        (p0.Z + p1.Z + p2.Z) / 3.0,
    )


def calc_landuse_areas_with_roads(
    doc,
    road_regions,
    landuse_parent="Landuse",
    z_height=1.0,
    z_limit=0.1,
    faces_collector=None,
):
    """
    Landuse 하위 레이어별 도로 영역을 빼고, z_limit 이하 면적 합산
    """

    def compute_total_area_for_landuse(
        lu_srfs, road_breps_list, z_h, z_lim, collected_faces_list=None
    ):
        total = 0.0
        before_area = sum(ghcomp.Area(srf).area for srf in lu_srfs)

        print("Before : ", before_area)
        decons = []
        for lu_srf in lu_srfs:
            # Landuse Extrude
            base_brep = extrude_srf(lu_srf, z_h)
            diff_brep = base_brep

            # 도로영역 차감

            for road_brep in road_breps_list:
                diff_brep = ghcomp.SolidDifference(diff_brep, road_brep)
                if not diff_brep:
                    break

            if not diff_brep:
                raise Exception("FAILED TO SUBTRACT ROADS")

            # Deconstruct Brep → Faces 추출
            decon = ghcomp.DeconstructBrep(diff_brep)
            decons.extend(decon[0])
            faces = decon[0]  # DeconstructBrep 결과는 (Faces, Edges, Vertices)

            if not faces:
                continue

            # Z 기준 필터링 후 면적 계산
            for f in faces:
                mp = rg.AreaMassProperties.Compute(f)
                if mp and mp.Centroid.Z > z_lim:
                    continue

                if not is_point_on_srf(get_point_inside_face(f), lu_srf):
                    continue

                total += mp.Area
                if collected_faces_list is not None:
                    collected_faces_list.append(f)

        if before_area < total:
            sc.sticky["test"] = decons
            raise Exception(
                "RoadSubError: Landuse area increased after road subtraction."
            )

        print("After : ", total)
        return total

    results = {}

    # Landuse dict: {용도명: [Brep, ...]}
    landuse_dict = get_layer_surfaces(doc, landuse_parent)

    # 도로 영역 Extrude
    road_breps = []
    for r in road_regions:
        road_brep = extrude_srf(r, z_height)
        if road_brep:
            road_breps.append(road_brep)

    # Helper: compute total area for a list of landuse surfaces after
    # subtracting roads and filtering by Z height.
    # 각 용도별 면적 계산 (helper 호출)

    print("Road Breps : ", len(road_breps))
    for lu_name, lu_srfs in landuse_dict.items():
        collected = [] if faces_collector is not None else None
        total_area = compute_total_area_for_landuse(
            lu_srfs, road_breps, z_height, z_limit, collected_faces_list=collected
        )

        results[lu_name] = total_area
        if faces_collector is not None:
            faces_collector[lu_name] = collected or []

    return results


# --------------------------------------------
# Excel Export Utilities (Windows + Excel 설치 필요)
# --------------------------------------------

# 템플릿 내 Landuse/Area/Percent 컬럼 탐색 시 사용할 기본 컬럼 인덱스 예비값 (D/E/F)
DEFAULT_COL_LANDUSE = 4
DEFAULT_COL_AREA = 5
DEFAULT_COL_PERC = 6


def get_site_boundary_area(doc, layer_name="SiteBoundary"):
    """SiteBoundary 레이어의 폐곡선 평면 커브 중 최대 면적을 사이트 총면적으로 사용."""
    objs = doc.Objects.FindByLayer(layer_name) or []
    if not objs:
        return None, 0.0

    tol = doc.ModelAbsoluteTolerance
    best_curve = None
    best_area = 0.0

    for obj in objs:
        crv = obj.Geometry if isinstance(obj.Geometry, rg.Curve) else None
        if crv is None:
            continue

        if not crv.IsClosed:
            continue

        best_curve = crv

        best_area = rg.AreaMassProperties.Compute(crv).Area

    return best_curve, best_area


def _build_output_excel_path(save_dir):
    """
    YYYYMMDD_AreaTable_num.xlsx 파일명을 save_dir 하위로 생성 (중복 시 num 증가)
    """
    if not save_dir:
        return None
    try:
        if not os.path.exists(save_dir):
            os.makedirs(save_dir)
    except Exception:
        # 디렉토리 생성 실패 시 그대로 진행 (뒤에서 실패처리)
        pass

    date_str = datetime.now().strftime("%Y%m%d")
    num = 1
    while True:
        name = f"{date_str}_AreaTable_{num}.xlsx"
        full = os.path.join(save_dir, name)
        if not os.path.exists(full):
            return full
        num += 1


def _ensure_unique_path(path):
    """경로가 존재하면 _1, _2 ... 를 파일명 뒤(확장자 앞)에 붙여 유니크 경로 반환"""
    if not os.path.exists(path):
        return path
    root, ext = os.path.splitext(path)
    i = 1
    while True:
        candidate = f"{root}_{i}{ext or '.xlsx'}"
        if not os.path.exists(candidate):
            return candidate
        i += 1


def _find_header_and_columns(ws):
    """
    템플릿 시트에서 'Landuse', 'Area', 'Perc' 컬럼을 찾고 헤더 행(row)을 반환.
    찾지 못하면 기본값(D/E/F) 사용.
    Returns: (header_row, col_landuse, col_area, col_perc)
    """
    try:
        used = ws.UsedRange
        start_row = used.Row
        start_col = used.Column
        row_count = used.Rows.Count
        col_count = used.Columns.Count
    except Exception:
        # UsedRange 접근 실패 시 안전한 기본값
        return (1, DEFAULT_COL_LANDUSE, DEFAULT_COL_AREA, DEFAULT_COL_PERC)

    header_row = None
    col_landuse = None
    col_area = None
    col_perc = None

    max_rows_to_scan = min(start_row + row_count, start_row + 30)
    max_cols_to_scan = min(start_col + col_count, start_col + 20)

    for r in range(start_row, max_rows_to_scan + 1):
        for c in range(start_col, max_cols_to_scan + 1):
            try:
                v = ws.Cells(r, c).Value2
            except Exception:
                v = None
            if isinstance(v, basestring):
                txt = v.strip().lower()
                if "landuse" in txt and col_landuse is None:
                    col_landuse = c
                    header_row = r
                if "area" in txt and col_area is None:
                    col_area = c
                if "perc" in txt or "%" in txt:
                    if col_perc is None:
                        col_perc = c

        # 기본적으로 첫 헤더 행만 찾으면 종료
        if header_row is not None:
            break

    if header_row is None:
        header_row = start_row
    if col_landuse is None:
        col_landuse = DEFAULT_COL_LANDUSE
    if col_area is None:
        col_area = DEFAULT_COL_AREA
    if col_perc is None:
        col_perc = DEFAULT_COL_PERC

    return (header_row, col_landuse, col_area, col_perc)


def _open_excel_app():
    excel_type = Type.GetTypeFromProgID("Excel.Application")
    if excel_type is None:
        raise Exception("Excel이 설치되어 있지 않거나 접근할 수 없습니다.")
    excel = Activator.CreateInstance(excel_type)
    # 가시성/경고 off
    try:
        excel.Visible = False
        excel.DisplayAlerts = False
    except Exception:
        pass
    return excel


def write_area_table_to_excel(template_path, output_path, landuse_areas):
    """
    템플릿 엑셀을 열어 Landuse 행 이름을 기준으로 면적/퍼센트를 채우고,
    저장 규칙에 맞게 새 파일로 저장.
    - template_path: 템플릿 파일 경로
    - output_path: 저장할 경로 (파일명 포함)
    - landuse_areas: {용도명: 면적}
    """
    if not template_path or not os.path.exists(template_path):
        raise Exception("template_path가 유효하지 않습니다.")
    if not output_path:
        raise Exception("output_path가 유효하지 않습니다.")

    total_area = sum(landuse_areas.values()) if landuse_areas else 0.0

    # 이름 매칭 추적
    remaining = dict(landuse_areas)

    excel = _open_excel_app()
    wb = None
    try:
        wb = excel.Workbooks.Open(template_path)
        ws = wb.Worksheets.Item(1)

        header_row, col_landuse, col_area, col_perc = _find_header_and_columns(ws)

        # UsedRange 기반으로 데이터 영역 순회하며 이름 매칭
        used = ws.UsedRange
        start_row = used.Row
        row_count = used.Rows.Count
        last_row = start_row + row_count + 50  # 약간 여유를 둔다

        total_row_index = None
        etc_row_index = None

        for r in range(header_row + 1, last_row + 1):
            try:
                name_val = ws.Cells(r, col_landuse).Value2
            except Exception:
                name_val = None

            if name_val is None:
                continue

            # 문자열화
            if not isinstance(name_val, basestring):
                try:
                    name_val = str(name_val)
                except Exception:
                    continue

            name = name_val.strip()
            name_lower = name.lower()

            if name_lower == "total":
                total_row_index = r
                continue

            if name_lower in ("etc", "기타"):
                etc_row_index = r
                continue

            if name in remaining:
                area_val = float(remaining[name])
                try:
                    ws.Cells(r, col_area).Value2 = area_val
                except Exception:
                    pass
                if total_area > 0.0:
                    try:
                        ws.Cells(r, col_perc).Value2 = (area_val / total_area) * 100.0
                    except Exception:
                        pass
                # 채워졌으면 제거
                del remaining[name]

        # 남은 항목은 Etc에 합산
        etc_area = sum(remaining.values()) if remaining else 0.0

        if etc_area > 0.0:
            r = etc_row_index
            if r is None:
                # 마지막 데이터 아래에 새로 추가
                r = last_row + 1
                try:
                    ws.Cells(r, col_landuse).Value2 = "Etc"
                except Exception:
                    pass
            try:
                ws.Cells(r, col_area).Value2 = float(etc_area)
            except Exception:
                pass
            if total_area > 0.0:
                try:
                    ws.Cells(r, col_perc).Value2 = (
                        float(etc_area) / total_area
                    ) * 100.0
                except Exception:
                    pass

        # Total 행 쓰기
        if total_row_index is not None:
            try:
                ws.Cells(total_row_index, col_area).Value2 = float(total_area)
            except Exception:
                pass
            try:
                ws.Cells(total_row_index, col_perc).Value2 = (
                    100.0 if total_area > 0 else 0.0
                )
            except Exception:
                pass

        # 새 파일로 저장
        wb.SaveAs(output_path)
    finally:
        try:
            if wb is not None:
                wb.Close(SaveChanges=False)
        except Exception:
            pass
        try:
            excel.Quit()
        except Exception:
            pass
        # COM 해제 (메모리 누수 방지)
        try:
            Marshal.ReleaseComObject(excel)
        except Exception:
            pass


def write_area_table_openpyxl(output_path, landuse_areas):
    """템플릿이 없거나 Excel COM을 사용할 수 없을 때, 간단한 표 형식으로 xlsx 생성."""
    if not OPENPYXL_AVAILABLE:
        raise Exception("openpyxl이 설치되어 있지 않아 새 파일을 만들 수 없습니다.")

    total = sum(landuse_areas.values()) if landuse_areas else 0.0
    rows = sorted(landuse_areas.items(), key=lambda kv: kv[1], reverse=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "Area Report"

    # 헤더
    ws.append(["Landuse", "Area (sq m)", "Percentage (%)"])

    # 본문
    for name, area in rows:
        pct = (area / total * 100.0) if total > 0 else 0.0
        ws.append([name, float(area), pct])

    # 합계
    ws.append([])
    ws.append(["Total", float(total), 100.0 if total > 0 else 0.0])

    # 간단 스타일
    header_font = Font(bold=True, name="Arial", size=11)
    center_align = Alignment(horizontal="center", vertical="center")
    right_align = Alignment(horizontal="right", vertical="center")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    for cell in ws[1]:
        cell.font = header_font
        cell.alignment = center_align
        cell.border = thin_border
    # 숫자/정렬
    for row in ws.iter_rows(min_row=2, min_col=2, max_col=3):
        for cell in row:
            cell.number_format = "#,##0.00"
            cell.alignment = right_align
            cell.border = thin_border

    # 자동 폭
    col_widths = {}
    for row in ws.iter_rows():
        for c in row:
            if c.value is None:
                continue
            w = len(str(c.value)) + 2
            col_widths[c.column] = max(col_widths.get(c.column, 8), w)
    for col, w in col_widths.items():
        ws.column_dimensions[get_column_letter(col)].width = w

    wb.save(output_path)


def build_report_text(landuse_areas):
    """면적표 텍스트 리포트 생성 (정렬/정렬폭 포함, 샘플 포맷 참고)"""
    if not landuse_areas:
        return "No areas computed."

    # 그대로 레이어(용도) 이름을 사용해 정렬 및 합계
    rows = sorted(landuse_areas.items(), key=lambda kv: kv[1], reverse=True)
    total = sum(val for _, val in rows)

    # 3) 서식화 (샘플 스타일)
    title = "--- Area Report ({}) ---".format(datetime.now().date())
    header = "{:<25} | {:>15} | {:>12}".format("Landuse", "Area (sq m)", "Percentage")
    sep = "-" * 57

    lines = [title, header, sep]
    for name, val in rows:
        perc = (val / total * 100.0) if total > 0 else 0.0
        lines.append("{:<25} | {:>15,.2f} | {:>10.2f} %".format(name, val, perc))

    lines.append(sep)
    lines.append(
        "{:<25} | {:>15,.2f} | {:>10.2f} %".format(
            "Total", total, 100.0 if total > 0 else 0.0
        )
    )

    return "\n".join(lines)


# --------------------------------------------
# 🧩 GH Python Inputs
# road_regions : List[Surface or Brep]
# --------------------------------------------
doc = Rhino.RhinoDoc.ActiveDoc
road_regions = globals().get("road_regions", [])
road_regions = list(road_regions)
print(road_regions)
run = globals().get("run", False)
site_layer = globals().get("site_layer", "SiteBoundary")
DEBUG = globals().get("debug", True)
# 새 기능: 도로가 잘려나온 랜드유즈 면을 베이크할지 여부 (GH Boolean)
BAKE_ROAD_SUBREGION = bool(globals().get("BakeRoadSubRegion", True))


def dprint(*args):
    if DEBUG:
        try:
            print(*args)
        except Exception:
            pass


if not run:
    raise Exception("run 입력이 False입니다. 실행을 원하면 True로 설정하세요.")

# Landuse 면적 계산: 이제 도로 차감 없이 순수 용도 면적만 계산
# (요청에 따라 도로는 총면적 - 용도합으로 계산)
# Landuse 면적 계산 + (선택) 베이크용 Face 동시 수집
# 주의: 수집 여부는 BAKE_ROAD_SUBREGION 플래그를 그대로 사용해 일관성 유지
_faces_for_bake = {} if BAKE_ROAD_SUBREGION else None
a = calc_landuse_areas_with_roads(
    doc,
    road_regions,
    landuse_parent="Landuse",
    z_height=1.0,
    z_limit=0.1,
    faces_collector=_faces_for_bake,
)

# 수집된 Face 키/개수 디버그 출력 (선택)
if _faces_for_bake is not None:
    try:
        dprint("Collected faces keys:", list(_faces_for_bake.keys()))
        dprint(
            "Collected counts:",
            {k: (len(v) if v is not None else 0) for k, v in _faces_for_bake.items()},
        )
    except Exception:
        pass


# -------------------------------------------------
# 선택 기능: BakeRoadSubRegion = True일 때, 'Landuse-Road' 트리에 결과 베이크
# -------------------------------------------------
def _find_layer_by_fullpath(doc, fullpath):
    for ly in doc.Layers:
        fp = getattr(ly, "FullPath", None) or ly.Name or ""
        if fp == fullpath:
            return ly
    return None


def _ensure_layer(doc, name, parent_id=None):
    # 동일 이름 + 동일 부모인 레이어 찾기
    if parent_id:
        parent = doc.Layers.FindId(parent_id)
        parent_path = getattr(parent, "FullPath", parent.Name) if parent else None
        target_path = (parent_path + "::" + name) if parent_path else name
        found = _find_layer_by_fullpath(doc, target_path)
        if found:
            return found
    else:
        # 최상위에서 이름 일치하는 첫 레이어 반환
        for ly in doc.Layers:
            if ly.ParentLayerId == System.Guid.Empty:
                if ly.Name == name:
                    return ly

    layer = Rhino.DocObjects.Layer()
    layer.Name = name
    if parent_id:
        layer.ParentLayerId = parent_id
        # 부모 레이어 색상 복제 (시각적 일관성)
        try:
            parent_layer = doc.Layers.FindId(parent_id)
            if parent_layer:
                layer.Color = parent_layer.Color
        except Exception:
            pass
    idx = doc.Layers.Add(layer)
    return doc.Layers[idx] if idx >= 0 else None


def _clear_layer_tree(doc, parent_name):
    # parent_name 하위 모든 레이어의 객체 삭제
    parent = None
    for ly in doc.Layers:
        if ly.Name == parent_name and (ly.ParentLayerId == System.Guid.Empty):
            parent = ly
            break
    if not parent:
        return 0
    parent_path = getattr(parent, "FullPath", None) or parent.Name or ""
    deleted = 0
    for ly in doc.Layers:
        fp = getattr(ly, "FullPath", None) or ly.Name or ""
        if parent_path and (fp == parent_path or fp.startswith(parent_path + "::")):
            objs = doc.Objects.FindByLayer(ly) or []
            for obj in objs:
                if doc.Objects.Delete(obj, True):
                    deleted += 1
    return deleted


def _duplicate_face_to_brep(face):
    try:
        return face.DuplicateFace(True)
    except Exception:
        try:
            srf = face.UnderlyingSurface()
            return rg.Brep.CreateFromSurface(srf)
        except Exception:
            return None


def _planar_breps_from_face(doc, face):
    """Return robust planar Breps from a BrepFace using duplicate or planar rebuild from loops."""
    breps = []
    fb = None
    try:
        fb = face.DuplicateFace(True)
        if fb and fb.IsValid:
            breps.append(fb)
    except Exception:
        fb = None
    if not breps:
        try:
            tol = doc.ModelAbsoluteTolerance if doc else 0.01
            crvs = []
            for loop in face.Loops:
                try:
                    crv = loop.To3dCurve()
                    if crv and crv.IsClosed:
                        crvs.append(crv)
                except Exception:
                    pass
            if crvs:
                made = rg.Brep.CreatePlanarBreps(crvs, tol)
                if made:
                    breps.extend([b for b in made if b and b.IsValid])
        except Exception:
            pass
    return breps


def compute_landuse_road_cut_faces_from_cache(doc, cached_face_dict, z_limit=0.1):
    """이미 면적 계산시 수집된 Face dict ({lu: [Face,...]}) 를 planar Brep로 변환.
    Returns { lu_name: [Brep,...] }"""
    result = {}
    if not cached_face_dict:
        return result
    for lu_name, faces in cached_face_dict.items():
        breps = []
        for f in faces:
            try:
                mp = rg.AreaMassProperties.Compute(f)
                if mp and mp.Centroid.Z > z_limit:
                    continue
            except Exception:
                pass
            fbreps = _planar_breps_from_face(doc, f)
            if fbreps:
                breps.extend(fbreps)
        result[lu_name] = breps
    return result


def bake_road_subregion_results(
    doc,
    road_regions,
    landuse_parent="Landuse",
    target_parent="Landuse-Road",
    z_height=1.0,
    z_limit=0.1,
    clear_existing=True,
):
    """
    'Landuse-Road' 부모 레이어를 만들고, 용도별 하위 레이어에 도로 차감 면을 베이크.
    Returns: baked object count
    """
    baked = 0
    # 레이어 준비
    parent_layer = _ensure_layer(doc, target_parent, parent_id=None)
    if not parent_layer:
        raise Exception("레이어 생성 실패: {}".format(target_parent))

    # 기존 결과 제거 (선택)
    if clear_existing:
        _clear_layer_tree(doc, target_parent)

    # 결과 계산
    # 캐시 사용: 면적 계산 단계에서 이미 추출한 Face 활용 (중복 연산 제거)
    faces_by_lu = (
        compute_landuse_road_cut_faces_from_cache(doc, _faces_for_bake, z_limit=z_limit)
        if _faces_for_bake is not None
        else {}
    )

    # 캐시가 비어있을 때의 안전장치: 필요 시 즉석 재계산으로 보정
    if (not faces_by_lu) and BAKE_ROAD_SUBREGION:
        try:
            dprint("[Bake] cache empty -> fallback recompute faces")
            _tmp_cache = {}
            calc_landuse_areas_with_roads(
                doc,
                road_regions,
                landuse_parent=landuse_parent,
                z_height=z_height,
                z_limit=z_limit,
                faces_collector=_tmp_cache,
            )
            faces_by_lu = compute_landuse_road_cut_faces_from_cache(
                doc, _tmp_cache, z_limit=z_limit
            )
            try:
                dprint(
                    "[Bake] fallback faces counts:",
                    {k: len(v) for k, v in faces_by_lu.items()},
                )
            except Exception:
                pass
        except Exception as _fb_err:
            dprint("[Bake] fallback recompute failed:", _fb_err)
    for lu_name, faces in faces_by_lu.items():
        # 디버그: 각 용도별 생성된 면 수
        try:
            dprint("[Bake] {} -> {} breps".format(lu_name, len(faces)))
        except Exception:
            pass
        child_layer = _ensure_layer(doc, lu_name, parent_id=parent_layer.Id)
        # 기존 Landuse 동일 이름 레이어 색상 가져오기
        try:
            original = _find_layer_by_fullpath(doc, "Landuse::" + lu_name)
            if original:
                lyr = doc.Layers[child_layer.Index]
                lyr.Color = original.Color
                try:
                    doc.Layers.Modify(lyr, child_layer.Index, True)
                except Exception:
                    pass
        except Exception:
            pass
        if not child_layer:
            continue
        layer_index = child_layer.Index
        attrs = Rhino.DocObjects.ObjectAttributes()
        attrs.LayerIndex = layer_index
        for fb in faces:
            try:
                if isinstance(fb, rg.Brep):
                    obj_id = doc.Objects.AddBrep(fb, attrs)
                elif isinstance(fb, rg.Surface):
                    obj_id = doc.Objects.AddSurface(fb, attrs)
                else:
                    # 마지막 시도: Brep 변환
                    brep_try = rg.Brep.TryConvertBrep(fb)
                    obj_id = (
                        doc.Objects.AddBrep(brep_try, attrs)
                        if brep_try
                        else System.Guid.Empty
                    )
                if obj_id and obj_id != System.Guid.Empty:
                    baked += 1
            except Exception:
                pass

    try:
        doc.Views.Redraw()
    except Exception:
        pass

    return baked


try:
    baked_cnt = bake_road_subregion_results(
        doc,
        road_regions,
        landuse_parent="Landuse",
        target_parent="Landuse-Road",
        z_height=1.0,
        z_limit=0.1,
        clear_existing=True,
    )
    dprint(
        "BakeRoadSubRegion: baked {} objects under 'Landuse-Road'.".format(baked_cnt)
    )
except Exception as _bake_err:
    dprint("BakeRoadSubRegion 오류:", _bake_err)

# --------------------------------------------
# GH Output: blocks (도로로 잘려나간 용도 블록 객체 리스트)
# - 우선 면적 계산 시 수집된 Face 캐시를 이용해 평면 Brep로 변환
# - 캐시가 없으면 안전하게 한 번 재계산하여 확보
# - 각 Brep를 Block(region, landuse_name, block_id) 객체로 포장하여 반환
# --------------------------------------------
blocks = []
try:
    _by_lu = (
        compute_landuse_road_cut_faces_from_cache(doc, _faces_for_bake, z_limit=0.1)
        if _faces_for_bake is not None
        else {}
    )
    if not _by_lu:
        _tmp_cache2 = {}
        calc_landuse_areas_with_roads(
            doc,
            road_regions,
            landuse_parent="Landuse",
            z_height=1.0,
            z_limit=0.1,
            faces_collector=_tmp_cache2,
        )
        _by_lu = compute_landuse_road_cut_faces_from_cache(
            doc, _tmp_cache2, z_limit=0.1
        )

    # Block 객체로 감싸서 평탄화하여 단일 리스트로 출력
    if _by_lu:
        per_lu_counts = {}
        for _lu, _lst in _by_lu.items():
            if not _lst:
                continue
            idx = per_lu_counts.get(_lu, 0)
            for _brep in _lst:
                idx += 1
                try:
                    bid = f"{_lu}_{idx}"
                    blocks.append(Block(region=_brep, landuse_name=_lu, block_id=bid))
                except Exception:
                    # 실패 시 스킵
                    pass
            per_lu_counts[_lu] = idx
    dprint("blocks count:", len(blocks))
except Exception as _blocks_err:
    dprint("blocks compute error:", _blocks_err)
    blocks = []

# 사이트 총면적 (SiteBoundary 최대 폐곡선)
site_curve, site_area = get_site_boundary_area(doc, site_layer)

# 도로 면적 = 총면적 - 용도합
sum_landuse = sum(a.values()) if a else 0.0
road_area = site_area - sum_landuse
warning_msg = None
if road_area < 0:
    warning_msg = "[WARN] Sum of landuse exceeds site area by {:.2f} m².".format(
        -road_area
    )
    road_area = 0.0

# 선택 입력 (GH에서 제공 가능): save_path(디렉토리), template_path, save_file(Boolean 토글)
save_path = globals().get("save_path", None)
template_path = globals().get("template_path", None)
save_file = globals().get("save_file", False)  # True면 저장 실행
try:
    save_flag = bool(save_file)
except Exception:
    save_flag = False

# 출력용 딕셔너리에 Road 반영
areas_for_output = dict(a)
areas_for_output["Road"] = road_area

# 엑셀로 저장
b = None  # GH 두 번째 출력: 저장된 파일 경로 또는 오류 메시지
report = build_report_text(areas_for_output)  # GH용 리포트 문자열

# 디버그 로그 시작
debug_lines = []
debug_lines.append("=== Area Export Debug ===")
debug_lines.append("Site layer: {}".format(site_layer))
debug_lines.append("Site area: {:,.2f}".format(site_area))
debug_lines.append("Landuse categories: {}".format(len(a)))
debug_lines.append("Sum landuse: {:,.2f}".format(sum_landuse))
debug_lines.append("Road area: {:,.2f}".format(road_area))
debug_lines.append("Save flag: {}".format(save_flag))
debug_lines.append("Save path: {}".format(save_path or "<none>"))
if warning_msg:
    debug_lines.append(warning_msg)
for ln in debug_lines:
    dprint(ln)
try:
    out_path = None
    if save_flag:
        if save_path:
            out_path = _build_output_excel_path(save_path)
        else:
            b = "저장 실패: save_path(디렉토리)를 지정하세요."
        dprint("Generated out_path:", out_path)

    if out_path and save_flag:
        # 1) 템플릿이 명시되었을 때: Excel COM -> 실패 시 openpyxl fallback
        if template_path and os.path.exists(template_path):
            dprint("Using template_path:", template_path)
            try:
                write_area_table_to_excel(template_path, out_path, areas_for_output)
                b = out_path
                dprint("Saved via Excel COM ->", b)
            except Exception as com_err:
                dprint("Excel COM save failed:", com_err)
                if OPENPYXL_AVAILABLE:
                    write_area_table_openpyxl(out_path, areas_for_output)
                    b = out_path
                    dprint("Fallback to openpyxl ->", b)
                else:
                    b = "Excel COM 오류: {} (openpyxl 미설치)".format(str(com_err))
                    dprint(b)
        else:
            # 2) 스크립트/저장 폴더에서 template(.xlsx) 탐색
            candidates = []
            base_dirs = [
                os.path.dirname(__file__) if "__file__" in globals() else os.getcwd(),
                os.path.dirname(out_path) if out_path else os.getcwd(),
            ]
            for d in base_dirs:
                for nm in ("template.xlsx", "Template.xlsx", "TEMPLATE.xlsx"):
                    candidates.append(os.path.join(d, nm))
            picked = next((p for p in candidates if os.path.exists(p)), None)
            dprint("Template candidates:", candidates)
            dprint("Picked template:", picked)

            if picked:
                try:
                    write_area_table_to_excel(picked, out_path, areas_for_output)
                    b = out_path
                    dprint("Saved via Excel COM (picked) ->", b)
                except Exception as com_err:
                    dprint("Excel COM save failed (picked):", com_err)
                    if OPENPYXL_AVAILABLE:
                        write_area_table_openpyxl(out_path, areas_for_output)
                        b = out_path
                        dprint("Fallback to openpyxl ->", b)
                    else:
                        b = "Excel COM 오류: {} (openpyxl 미설치)".format(str(com_err))
                        dprint(b)
            else:
                # 3) 템플릿이 전혀 없으면 새 파일 생성 (openpyxl 필요)
                if OPENPYXL_AVAILABLE:
                    write_area_table_openpyxl(out_path, areas_for_output)
                    b = out_path
                    dprint("Saved via openpyxl (no template) ->", b)
                else:
                    b = "template.xlsx를 찾을 수 없고 openpyxl도 없습니다. template_path를 제공하거나 openpyxl을 설치하세요."
                    dprint(b)
    # save_flag=False이면 저장을 시도하지 않음 (b는 None 유지)
except Exception as _excel_err:
    try:
        b = str(_excel_err)
    except Exception:
        b = "엑셀 저장 중 알 수 없는 오류 발생"
    dprint("Save exception:", b)

# 최종 out 텍스트 구성 및 출력
out_lines = []
out_lines.append("Site: {:,.2f} ㎡".format(site_area))
out_lines.append("Total Landuse: {:,.2f} ㎡".format(sum_landuse))
out_lines.append("Road: {:,.2f} ㎡".format(road_area))
if warning_msg:
    out_lines.append(warning_msg)
if save_flag:
    out_lines.append("Saved: {}".format(b if b else "<no path>"))
else:
    out_lines.append("Save skipped (save_file=False)")
out = "\n".join(out_lines)
dprint("\n=== OUT ===\n" + out)
