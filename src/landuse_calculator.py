import os
from datetime import datetime

import Rhino
import Rhino.Geometry as rg


try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, Side
    from openpyxl.utils import get_column_letter

    OPENPYXL_AVAILABLE = True
except Exception:
    OPENPYXL_AVAILABLE = False


def collect_landuse_road_breps(doc, parent_name="Landuse-Road"):
    """Return {layer_name: [Brep, ...]} for every child under parent_name."""
    if doc is None:
        return {}

    results = {}
    target_prefix = parent_name + "::"

    for layer in doc.Layers:
        full_path = getattr(layer, "FullPath", layer.Name) or ""
        if full_path == parent_name:
            continue
        if not full_path.startswith(target_prefix):
            continue

        parts = full_path.split("::")
        if len(parts) < 2:
            continue
        child_name = parts[1]

        objs = doc.Objects.FindByLayer(layer) or []
        breps = results.setdefault(child_name, [])
        for obj in objs:
            geo = obj.Geometry
            if isinstance(geo, rg.Brep):
                breps.append(geo)
            elif isinstance(geo, rg.Surface):
                brep = rg.Brep.CreateFromSurface(geo)
                if brep:
                    breps.append(brep)

    return results


def compute_area_by_layer(layer_breps):
    """Compute summed planar area for each layer name."""
    areas = {}
    for layer_name, breps in (layer_breps or {}).items():
        total = 0.0
        for brep in breps:
            try:
                mp = rg.AreaMassProperties.Compute(brep)
                total += mp.Area if mp else 0.0
            except Exception:
                continue
        areas[layer_name] = total
    return areas


def get_single_site_boundary_curve(doc, layer_name="SiteBoundary"):
    if doc is None:
        raise Exception("Rhino document is not available.")

    objs = doc.Objects.FindByLayer(layer_name) or []
    closed_curves = []
    for obj in objs:
        geo = obj.Geometry
        if isinstance(geo, rg.Curve) and geo.IsClosed:
            closed_curves.append(geo)

    if len(closed_curves) != 1:
        raise Exception(
            "Layer '{}' must contain exactly one closed curve (found {}).".format(
                layer_name, len(closed_curves)
            )
        )

    return closed_curves[0]


def compute_curve_area(curve):
    mp = rg.AreaMassProperties.Compute(curve)
    if not mp:
        raise Exception("Failed to compute SiteBoundary area.")
    return mp.Area


def build_table_rows(area_table, site_area=None, road_area=None):
    area_table = area_table or {}
    rows = sorted(area_table.items(), key=lambda item: item[1], reverse=True)
    landuse_total = sum(area for _, area in rows)

    total_area = site_area
    if total_area is None:
        total_area = landuse_total + (road_area if road_area is not None else 0.0)

    percent_base = total_area if total_area and total_area > 0 else None

    def as_percent(val):
        if percent_base and percent_base > 0:
            return (val / percent_base) * 100.0
        return 0.0

    table = []
    for name, val in rows:
        table.append((name, val, as_percent(val)))

    if road_area is not None:
        table.append(("Road Area", road_area, as_percent(road_area)))

    table.append(("Total", total_area, 100.0 if percent_base else 0.0))

    return table


def build_report_text(area_table, site_area=None, road_area=None):
    lines = []
    table_rows = build_table_rows(area_table, site_area, road_area)
    if not table_rows:
        return "No Landuse-Road layers found."

    title = "--- Landuse-Road Area Report ({}) ---".format(datetime.now().date())
    header = "{:<25} | {:>15} | {:>12}".format("Layer", "Area (㎡)", "Percent")
    sep = "-" * 57

    lines.extend([title, header, sep])
    for name, val, pct in table_rows:
        if name == "Total":
            lines.append(sep)
        lines.append("{:<25} | {:>15,.2f} | {:>10.2f} %".format(name, val, pct))

    return "\n".join(lines)


def ensure_save_directory(path):
    if not path:
        return None
    try:
        os.makedirs(path, exist_ok=True)
    except Exception:
        return None
    return path


def next_report_path(save_dir):
    if not save_dir:
        return None
    date_prefix = datetime.now().strftime("%Y%m%d")
    idx = 1
    while True:
        candidate = os.path.join(save_dir, f"{date_prefix}_LanduseRoad_{idx}.xlsx")
        if not os.path.exists(candidate):
            return candidate
        idx += 1


def write_openpyxl_report(path, area_table, site_area=None, road_area=None):
    if not OPENPYXL_AVAILABLE:
        raise Exception("openpyxl not installed; install it or disable save_file.")

    wb = Workbook()
    ws = wb.active
    ws.title = "Landuse-Road"

    ws.append(["Layer", "Area (㎡)", "Percent (%)"])
    for name, area, pct in build_table_rows(area_table, site_area, road_area):
        ws.append([name, float(area), pct])

    header_font = Font(bold=True)
    right_align = Alignment(horizontal="right")
    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for cell in ws[1]:
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
        cell.border = border

    for row in ws.iter_rows(min_row=2, min_col=2, max_col=3):
        for cell in row:
            cell.number_format = "#,##0.00"
            cell.alignment = right_align
            cell.border = border

    widths = {}
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is None:
                continue
            w = len(str(cell.value)) + 2
            idx = cell.column
            widths[idx] = max(widths.get(idx, 0), w)

    for col, width in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = width

    wb.save(path)


def save_area_table(save_flag, save_dir, area_table, site_area=None, road_area=None):
    if not save_flag:
        return None

    target_dir = ensure_save_directory(save_dir)
    if not target_dir:
        return "Failed to create save directory."

    if not area_table and site_area is None and road_area is None:
        return "No data to save."

    out_path = next_report_path(target_dir)
    try:
        write_openpyxl_report(out_path, area_table, site_area, road_area)
        return out_path
    except Exception as err:
        return str(err)


def format_out_text(area_table, saved_path, save_flag, site_area=None, road_area=None):
    lines = []
    total = sum(area_table.values()) if area_table else 0.0
    lines.append("Landuse-Road layers: {}".format(len(area_table)))
    lines.append("Total area: {:,.2f} ㎡".format(total))
    if site_area is not None:
        lines.append("SiteBoundary area: {:,.2f} ㎡".format(site_area))
    if road_area is not None:
        lines.append("Road area: {:,.2f} ㎡".format(road_area))
    if save_flag:
        lines.append("Saved: {}".format(saved_path or "<error>"))
    else:
        lines.append("Save skipped (save_file=False)")
    return "\n".join(lines)


doc = Rhino.RhinoDoc.ActiveDoc
run = bool(globals().get("run", False))
save_file = bool(globals().get("save_file", False))
save_path = globals().get("save_path", None)
site_layer = globals().get("site_layer", "SiteBoundary")

if not run:
    a = {}
    b = None
    report = "Run flag is False. Set run=True to execute."
    out = "Save skipped (component idle)."
    blocks = []
else:
    breps_by_layer = collect_landuse_road_breps(doc)
    a = compute_area_by_layer(breps_by_layer)
    site_curve = get_single_site_boundary_curve(doc, site_layer)
    site_area = compute_curve_area(site_curve)
    landuse_total = sum(a.values()) if a else 0.0
    road_area = site_area - landuse_total
    warning = None
    if road_area < 0:
        warning = (
            "Warning: Landuse area exceeds SiteBoundary area by {:,.2f} ㎡.".format(
                abs(road_area)
            )
        )
    b = save_area_table(save_file, save_path, a, site_area, road_area)
    report = build_report_text(a, site_area, road_area)
    out = format_out_text(a, b, save_file, site_area, road_area)
    if warning:
        out = out + "\n" + warning
        report = report + "\n" + warning
    blocks = []
